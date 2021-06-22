import argparse
from openpyxl import load_workbook
import logging

### Set up the logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
s_handler = logging.StreamHandler()
s_formatter = logging.Formatter('%(message)s')
s_handler.setFormatter(s_formatter)
logger.addHandler(s_handler)

### Start of processing
logger.info("Starting the update")

### Get the filename from the command line and open the file for processing
parser = argparse.ArgumentParser(description='Update knock-out stage bracket with team names.', usage='knockout_stage_updater.py <file_name> <list_of_replacements>')
parser.add_argument('file_name', help='The name of the file to process')
parser.add_argument('list_of_replacements', help='A string representing a list of names to replace in the knock-out bracket in the format (<name_to_replace>, <replacement_name>, ...)')
args = parser.parse_args()

filename = args.file_name
list_of_replacements = args.list_of_replacements.split(",")
logger.info("Updating file {}".format(filename))

replacement_names = {}
for i in range(0, len(list_of_replacements), 2):
    replacement_names[list_of_replacements[i].strip()] = list_of_replacements[i+1].strip()

file = None

try:
    file = load_workbook(filename=filename)
except FileNotFoundError:
    raise SystemExit("No such file or directory: '{}'".format(filename))

sheetnames = file.sheetnames
for sheetname in sheetnames:
    if sheetname in ('Leaderboard'):
        logger.debug("'{}' not a knock-out stage sheet - continuing.".format(sheetname))
        continue
    
    logger.info("Updating {}".format(sheetname))
    for col in file[sheetname].iter_cols(min_row=43, max_row=65, min_col=3, max_col=15):
        for cell in col:
            if replacement_names.get(cell.value):
                logger.info("    Replacing {} with {}".format(cell.value, replacement_names[cell.value]))
                cell.value = replacement_names[cell.value]
            
file.save(filename=filename)
